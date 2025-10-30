"""Document parsing service for Word and Excel files."""
from __future__ import annotations

import io
from typing import Any, Dict, List, Optional

try:
    from docx import Document
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class DocumentParser:
    """Parses Word (.docx) and Excel (.xlsx) documents."""

    def __init__(self):
        """Initialize parser."""
        self.max_file_size = 50 * 1024 * 1024  # 50MB limit

    async def parse_word_document(self, buffer: bytes) -> str:
        """Extract text from Word document.

        Args:
            buffer: File buffer bytes

        Returns:
            Extracted text content
        """
        if not HAS_DOCX:
            raise RuntimeError("python-docx is not installed")

        if len(buffer) > self.max_file_size:
            raise ValueError(f"File size exceeds limit of {self.max_file_size}")

        try:
            doc = Document(io.BytesIO(buffer))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n".join(paragraphs)
        except Exception as e:
            raise ValueError(f"Failed to parse Word document: {str(e)}")

    async def parse_excel(self, buffer: bytes) -> List[List[str]]:
        """Parse Excel spreadsheet into rows.

        Args:
            buffer: File buffer bytes

        Returns:
            List of rows (each row is a list of cell values)
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is not installed")

        if len(buffer) > self.max_file_size:
            raise ValueError(f"File size exceeds limit of {self.max_file_size}")

        try:
            wb = load_workbook(io.BytesIO(buffer), data_only=True)
            ws = wb.active
            rows: List[List[str]] = []

            for row in ws.iter_rows(values_only=True):
                row_data = [str(cell or "") for cell in row]
                rows.append(row_data)

            return rows
        except Exception as e:
            raise ValueError(f"Failed to parse Excel file: {str(e)}")

    async def extract_structured_content(self, buffer: bytes, file_type: str = "docx") -> Dict[str, str]:
        """Extract structured content from document.

        Args:
            buffer: File buffer bytes
            file_type: 'docx' or 'xlsx'

        Returns:
            Dict with 'html' and 'text' keys
        """
        if file_type == "docx":
            text = await self.parse_word_document(buffer)
            return {
                "html": self._text_to_html(text),
                "text": text,
                "format": "word",
            }
        elif file_type == "xlsx":
            rows = await self.parse_excel(buffer)
            text = self._rows_to_text(rows)
            return {
                "html": self._rows_to_html(rows),
                "text": text,
                "format": "excel",
            }
        else:
            raise ValueError(f"Unsupported file type: {file_type}")

    def _text_to_html(self, text: str) -> str:
        """Convert plain text to simple HTML."""
        paragraphs = text.split("\n")
        html_parts = ["<div class='document'>"]

        for para in paragraphs:
            para = para.strip()
            if para:
                html_parts.append(f"<p>{self._escape_html(para)}</p>")

        html_parts.append("</div>")
        return "\n".join(html_parts)

    def _rows_to_text(self, rows: List[List[str]]) -> str:
        """Convert rows to plain text."""
        lines = []
        for row in rows:
            lines.append("\t".join(row))
        return "\n".join(lines)

    def _rows_to_html(self, rows: List[List[str]]) -> str:
        """Convert rows to HTML table."""
        html_parts = ["<table border='1'>"]

        for row in rows:
            html_parts.append("<tr>")
            for cell in row:
                html_parts.append(f"<td>{self._escape_html(cell)}</td>")
            html_parts.append("</tr>")

        html_parts.append("</table>")
        return "\n".join(html_parts)

    @staticmethod
    def _escape_html(text: str) -> str:
        """Escape HTML special characters."""
        return (
            text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
            .replace("'", "&#39;")
        )
